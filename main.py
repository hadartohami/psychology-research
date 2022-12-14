import openpyxl

FILE_PATH = './analysis.xlsx'
WB_OBJ = openpyxl.load_workbook(FILE_PATH)
SHEET_OBJ = WB_OBJ.active
NOT_COMPLETED = "not completed"


def sh_scoring():
    rows = SHEET_OBJ.iter_rows(min_row=2, max_row=SHEET_OBJ.max_row, min_col=10, max_col=13)
    scores = []
    for a, b, c, d in rows:
        score = 0
        if a.value is not None:
            score += a.value
        if b.value is not None:
            score += b.value
        if c.value is not None:
            score += c.value
        if d.value is None:
            score = NOT_COMPLETED
            scores.append(score)
            continue
        elif d.value is not None:
            score += (8 - d.value)
        score = score / 4
        scores.append(score)
    score_column = 71
    SHEET_OBJ.cell(column=score_column, row=1, value="SHS Score")
    add_scores_to_file(scores, score_column)


def dass_general_score():
    # excel columns:
    columns = [73, 74, 75]
    scores = calculate_score(columns, check_one_empty_item=True)
    score_column = 72
    SHEET_OBJ.cell(column=score_column, row=1, value="Dass - General Score")
    add_scores_to_file(scores, score_column)


def dass_depression_score():
    # excel columns:
    columns = [16, 18, 23, 26, 29, 30, 34]
    scores = calculate_score(columns, check_one_empty_item=True)
    score_column = 73
    SHEET_OBJ.cell(column=score_column, row=1, value="Dass - Depression Score")
    add_scores_to_file(scores, score_column)


def dass_anxiety_score():
    # excel columns:
    columns = [15, 17, 20, 22, 28, 32, 33]
    score_column = 74
    SHEET_OBJ.cell(column=score_column, row=1, value="Dass - Anxiety Score")
    scores = calculate_score(columns, check_one_empty_item=True)
    add_scores_to_file(scores, score_column)


def dass_pressure_score():
    # excel columns:
    columns = [14, 19, 21, 24, 25, 27, 31]
    score_column = 75
    SHEET_OBJ.cell(column=score_column, row=1, value="Dass - Pressure Score")
    scores = calculate_score(columns, check_one_empty_item=True)
    add_scores_to_file(scores, score_column)


def ptsd_general_score():
    # excel columns: AX:BQ
    scores = []
    criterion_b_column = 77
    criterion_c_column = 79
    criterion_d_column = 81
    criterion_e_column = 83
    for r in range(2, SHEET_OBJ.max_row+1):
        score = 0
        cell_obj_b = SHEET_OBJ.cell(row=r, column=criterion_b_column)
        cell_obj_c = SHEET_OBJ.cell(row=r, column=criterion_c_column)
        cell_obj_d = SHEET_OBJ.cell(row=r, column=criterion_d_column)
        cell_obj_e = SHEET_OBJ.cell(row=r, column=criterion_e_column)
        if cell_obj_b.value is not None and cell_obj_c is not None and cell_obj_d.value is not None and cell_obj_e.value is not None:
            if cell_obj_b.value >= 1 and cell_obj_c.value >= 1 and cell_obj_d.value >= 2 and cell_obj_e.value >= 2:
                score = 1
        scores.append(score)
    score_column = 76
    SHEET_OBJ.cell(column=score_column, row=1, value="General PTSD Score")
    add_scores_to_file(scores, score_column)


def criterion_b_score():
    # excel columns: AX:BB
    columns = list(range(50, 54 + 1))
    scores = calculate_pcl_ptsd_score(columns)
    score_column = 77
    SHEET_OBJ.cell(column=score_column, row=1, value="Criterion B Score - Re-experiencing")
    add_scores_to_file(scores, score_column)


def criterion_b_percentage_score():
    column = 77
    num_of_questions = 5
    scores = calculate_criterion_ptsd_percentage_score(column, num_of_questions)
    score_column = 78
    SHEET_OBJ.cell(column=score_column, row=1, value="Criterion B Score In Percentage - Re-experiencing")
    add_scores_to_file(scores, score_column)


def criterion_c_score():
    # excel columns: BC:BD
    columns = list(range(55, 56 + 1))
    scores = calculate_pcl_ptsd_score(columns)
    score_column = 79
    SHEET_OBJ.cell(column=score_column, row=1, value="Criterion C Score - Avoidance")
    add_scores_to_file(scores, score_column)


def criterion_c_percentage_score():
    column = 79
    num_of_questions = 2
    scores = calculate_criterion_ptsd_percentage_score(column, num_of_questions)
    score_column = 80
    SHEET_OBJ.cell(column=score_column, row=1, value="Criterion C Score In Percentage - Avoidance")
    add_scores_to_file(scores, score_column)


def criterion_d_score():
    # excel columns: BE:BK
    columns = list(range(57, 63 + 1))
    scores = calculate_pcl_ptsd_score(columns)
    score_column = 81
    SHEET_OBJ.cell(column=score_column, row=1, value="Criterion D Score - Negative alterations in cognition & mood")
    add_scores_to_file(scores, score_column)


def criterion_d_percentage_score():
    column = 81
    num_of_questions = 7
    scores = calculate_criterion_ptsd_percentage_score(column, num_of_questions)
    score_column = 82
    SHEET_OBJ.cell(column=score_column, row=1, value="Criterion D Score In Percentage - Negative alterations in cognition & mood")
    add_scores_to_file(scores, score_column)


def criterion_e_score():
    # excel columns: BL:BQ
    columns = list(range(64, 69 + 1))
    scores = calculate_pcl_ptsd_score(columns)
    score_column = 83
    SHEET_OBJ.cell(column=score_column, row=1, value="Criterion E Score -  Hyper-arousal")
    add_scores_to_file(scores, score_column)


def criterion_e_percentage_score():
    column = 83
    num_of_questions = 6
    scores = calculate_criterion_ptsd_percentage_score(column, num_of_questions)
    score_column = 84
    SHEET_OBJ.cell(column=score_column, row=1, value="Criterion E Score In Percentage-  Hyper-arousal")
    add_scores_to_file(scores, score_column)


def calculate_criterion_ptsd_percentage_score(column, num_of_questions):
    scores = []
    for r in range(2, SHEET_OBJ.max_row+1):
        cell_obj = SHEET_OBJ.cell(row=r, column=column)
        score = round(((cell_obj.value / num_of_questions) * 100), 2 )
        scores.append(score)
    return scores


def calculate_score(columns, check_last_item=False, check_one_empty_item=False):
    # if last time is empty the score is "not_completed"
    # check_last_item = False
    # if one item is empty the score  is "not_completed"
    # check_one_empty_item = False
    scores = []
    break_out_flag = False
    for r in range(2, SHEET_OBJ.max_row+1):
        score = 0
        break_out_flag = False
        for c in columns:
            cell_obj = SHEET_OBJ.cell(row=r, column=c)
            if check_one_empty_item and (cell_obj.value is None or cell_obj.value == NOT_COMPLETED):
                break_out_flag = True
                score = NOT_COMPLETED
            if cell_obj.value is not None and isinstance(cell_obj.value, int):
                score = score + cell_obj.value
            if break_out_flag:
                break
            if check_last_item and c == columns[len(columns)-1] and cell_obj.value is None:
                score = NOT_COMPLETED
        if score == 0 or score == NOT_COMPLETED:
            score = NOT_COMPLETED
        scores.append(score)
    return scores


def calculate_pcl_criterion_score(columns):
    scores = []
    for r in range(2, SHEET_OBJ.max_row+1):
        score = 0
        for c in columns:
            cell_obj = SHEET_OBJ.cell(row=r, column=c)
            if cell_obj.value is not None and isinstance(cell_obj.value, int) and cell_obj.value >= 2:
                score = score + 1
        scores.append(score)
    return scores


def calculate_pcl_ptsd_score(columns):
    scores = []
    for r in range(2, SHEET_OBJ.max_row+1):
        score = 0
        for c in columns:
            cell_obj = SHEET_OBJ.cell(row=r, column=c)
            if cell_obj.value is not None and isinstance(cell_obj.value, int) and cell_obj.value >= 2:
                score = score + 1
        scores.append(score)
    return scores


def add_scores_to_file(scores, col):
    for i in range(len(scores)):
        insert_value(i+2, col, scores[i])


def insert_value(row, col, value):
    SHEET_OBJ.cell(column=col, row=row, value=value)


def main():
    SHEET_OBJ.insert_cols(72, amount=9)
    sh_scoring()
    dass_depression_score()
    dass_anxiety_score()
    dass_pressure_score()
    dass_general_score()
    criterion_b_score()
    criterion_b_percentage_score()
    criterion_c_score()
    criterion_c_percentage_score()
    criterion_d_score()
    criterion_d_percentage_score()
    criterion_e_score()
    criterion_e_percentage_score()
    ptsd_general_score()
    WB_OBJ.save(filename="scores.xlsx")


if __name__ == "__main__":
    main()
