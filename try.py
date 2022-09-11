import openpyxl

FILE_PATH = './test.xlsx'
WB_OBJ = openpyxl.load_workbook(FILE_PATH)
SHEET_OBJ = WB_OBJ.active


def sh_scoring():
    rows = SHEET_OBJ.iter_rows(min_row=2, max_row=SHEET_OBJ.max_row, min_col=10, max_col=13)
    scores = []
    d_is_empty = False
    for a, b, c, d in rows:
        score = 0
        if a.value is not None:
            score += a.value
        if b.value is not None:
            score += b.value
        if c.value is not None:
            score += c.value
        if d.value is None:
            score = "not completed"
            scores.append(score)
            continue
        elif d.value is not None:
            score += (8 - d.value)
        score = score / 4
        scores.append(score)
    score_column = 71
    SHEET_OBJ.cell(column=score_column, row=1, value="SHS Score")
    add_scores_to_file(scores, score_column)


def dass_depression_score():
    # excel columns:
    columns = [16, 18, 23, 26, 29, 30, 34]
    scores = calculate_score(columns, check_one_empty_item=True)
    score_column = 72
    SHEET_OBJ.cell(column=score_column, row=1, value="Depression Score")
    add_scores_to_file(scores, score_column)


def dass_anxiety_score():
    # excel columns:
    columns = [15, 17, 20, 22, 28, 32, 33]
    score_column = 73
    SHEET_OBJ.cell(column=score_column, row=1, value="Anxiety Score")
    scores = calculate_score(columns)
    add_scores_to_file(scores, score_column)


def dass_pressure_score():
    # excel columns:
    columns = [14, 19, 21, 24, 25, 27, 31]
    score_column = 74
    SHEET_OBJ.cell(column=score_column, row=1, value="Pressure Score")
    scores = calculate_score(columns)
    add_scores_to_file(scores, score_column)


def PTSD_general_score():
    # excel columns: AX:BQ
    columns = list(range(50, 69 + 1))
    scores = calculate_score(columns)
    score_column = 75
    SHEET_OBJ.cell(column=score_column, row=1, value="General PTSD Score")
    add_scores_to_file(scores, score_column)


def criterion_B_score():
    # excel columns: AX:BB
    columns = list(range(50, 54 + 1))
    scores = calculate_score(columns)
    score_column = 76
    SHEET_OBJ.cell(column=score_column, row=1, value="Criterion B Score - Re-experiencing")
    add_scores_to_file(scores, score_column)


def criterion_C_score():
    # excel columns: BC:BD
    columns = list(range(55, 56 + 1))
    scores = calculate_score(columns)
    score_column = 77
    SHEET_OBJ.cell(column=score_column, row=1, value="Criterion C Score - Avoidance")
    add_scores_to_file(scores, score_column)


def criterion_D_score():
    # excel columns: BE:BK
    columns = list(range(57, 63 + 1))
    scores = calculate_score(columns)
    score_column = 78
    SHEET_OBJ.cell(column=score_column, row=1, value="Criterion D Score - Negative alterations in cognition & mood")
    add_scores_to_file(scores, score_column)


def criterion_E_score():
    # excel columns: BL:BQ
    columns = list(range(64, 69 + 1))
    scores = calculate_score(columns)
    score_column = 79
    SHEET_OBJ.cell(column=score_column, row=1, value="Criterion E Score -  Hyper-arousal")
    add_scores_to_file(scores, score_column)


def calculate_score(columns, check_last_item=False, check_one_empty_item=False):
    # if last time is empty the score is "not_completed"
    # check_last_item = False
    # if one item is empty the score  is "not_completed"
    # check_one_empty_item = False
    scores = []
    for r in range(2, SHEET_OBJ.max_row+1):
        score = 0
        for c in columns:
            cell_obj = SHEET_OBJ.cell(row=r, column=c)
            if check_one_empty_item and cell_obj.value is None:
                score = "not completed"
                continue
            if cell_obj.value is not None and isinstance(cell_obj.value, int):
                score = score + cell_obj.value
        if score == 0 or score == "not completed":
            score = "not completed"
        scores.append(score)
    return scores


def add_scores_to_file(scores, col):
    for i in range(len(scores)):
        insert_value(i+2, col, scores[i])


def insert_value(row, col, value):
    SHEET_OBJ.cell(column=col, row=row, value=value)


def main():
    SHEET_OBJ.insert_cols(72, amount=9)
    sh_scoring()
    dass_depression_score()
    dass_anxiety_score()
    dass_pressure_score()
    PTSD_general_score()
    criterion_B_score()
    criterion_C_score()
    criterion_D_score()
    criterion_E_score()
    WB_OBJ.save(filename="scores.xlsx")


if __name__ == "__main__":
    main()
