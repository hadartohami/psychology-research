import openpyxl
import hashlib

FILE_PATH = './id.xlsx'
WB_OBJ = openpyxl.load_workbook(FILE_PATH)
SHEET_OBJ = WB_OBJ.active
ID_COLUMN = 2


def insert_value(row, col, value):
    SHEET_OBJ.cell(column=col, row=row, value=value)


def main():
    for row in range(2, SHEET_OBJ.max_row + 1):
        cell_obj = SHEET_OBJ.cell(row=row, column=ID_COLUMN)
        value = str(cell_obj.value).encode("ASCII")
        anonymised_id = hashlib.sha256(value).hexdigest()
        SHEET_OBJ.cell(column=ID_COLUMN, row=row, value=anonymised_id)
    WB_OBJ.save(filename="anonymised_data.xlsx")


if __name__ == "__main__":
    main()
